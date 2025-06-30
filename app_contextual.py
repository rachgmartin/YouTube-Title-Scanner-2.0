
import os
import base64
from dotenv import load_dotenv
import streamlit as st
import pandas as pd
import requests
from scan_titles_weighted_contextual_v3_riskaware import scan_titles_weighted

# Load variables from .env if present
load_dotenv()


st.set_page_config(page_title="YouTube Video Title Scanner", layout="centered")

with open("assets/Studio71_Icon_White.png", "rb") as img_file:
    logo_b64 = base64.b64encode(img_file.read()).decode()

st.markdown(
    f"""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@300;800&display=swap');
        body {
            background-color: #0E1117;
            color: #FFFFFF;
        }
        .branded-content {{
            font-family: 'Montserrat', sans-serif;
        }}
        .wave-section {{
            background: linear-gradient(135deg, #FF0230, #62A362);
            color: #FFFFFF;
            text-align: center;
            position: relative;
            padding: 3rem 1rem 6rem;
            overflow: hidden;
        }}
        .wave-section .logo {{
            width: 120px;
            margin-bottom: 1rem;
        }}
        .wave-section h1 {{
            font-weight: 800;
            margin-bottom: 0.2rem;
        }}
        .wave-section .subtitle {{
            font-weight: 300;
            margin: 0;
        }}
        .wave-section svg {{
            position: absolute;
            bottom: 0;
            left: 0;
            width: 100%;
            height: 80px;
        }}
        .results-table {{
        }}
    </style>
    <div class="branded-content wave-section">
        <img src="data:image/png;base64,{logo_b64}" class="logo" alt="Studio71 logo" />
        <h1>YouTube Video Title Scanner</h1>
        <p class="subtitle">Scan a YouTube channel for advertiser-unfriendly words and suggest safer alternatives.</p>
        <svg viewBox="0 0 1440 100" preserveAspectRatio="none">
            <path d="M0,30 C200,80 400,0 600,30 C800,60 1000,20 1200,40 C1300,50 1400,40 1440,30 L1440,100 L0,100 Z" fill="#0E1117"></path>
        </svg>
    </div>
    """,
    unsafe_allow_html=True,
)

api_key = st.text_input(
    "Enter your YouTube Data API Key",
    value=os.getenv("YOUTUBE_API_KEY", ""),
    type="password",
)
channel_id = st.text_input("Enter the YouTube Channel ID (e.g., UC_x5XG1OV2P6uZZ5FSM9Ttw)")
max_results = st.number_input("Maximum number of titles to fetch", min_value=1, max_value=500, value=100)

def get_uploads_playlist_id(api_key, channel_id):
    url = f"https://www.googleapis.com/youtube/v3/channels?part=contentDetails&id={channel_id}&key={api_key}"
    response = requests.get(url).json()
    try:
        return response['items'][0]['contentDetails']['relatedPlaylists']['uploads']
    except (KeyError, IndexError):
        return None

def fetch_video_titles(api_key, uploads_playlist_id, max_results):
    titles = []
    page_token = ""
    while len(titles) < max_results:
        url = f"https://www.googleapis.com/youtube/v3/playlistItems?key={api_key}&playlistId={uploads_playlist_id}&part=snippet&maxResults=50&pageToken={page_token}"
        response = requests.get(url).json()
        for item in response.get("items", []):
            titles.append(item["snippet"]["title"])
        page_token = response.get("nextPageToken", "")
        if not page_token:
            break
    return titles[:max_results]

if st.button("Scan Titles") and api_key and channel_id:
    try:
        st.info("Fetching video titles...")
        uploads_playlist_id = get_uploads_playlist_id(api_key, channel_id)
        if not uploads_playlist_id:
            st.error("Failed to retrieve uploads playlist. Check Channel ID.")
        else:
            titles = fetch_video_titles(api_key, uploads_playlist_id, max_results)
            if not titles:
                st.warning("No titles found or API quota exceeded.")
            else:
                df_keywords = pd.read_csv("updated_keywords_expanded.csv")
                df_severity = pd.read_csv("safety_severity_scores.csv")
                df_results = scan_titles_weighted(titles, df_keywords, df_severity)
                df_results = df_results.sort_values(by="Safety Score")
                styled_df = df_results.style.background_gradient(
                    cmap="RdYlGn", subset=["Safety Score"]
                )
                st.success("Scan complete!")
                st.markdown("<div class='results-table'>", unsafe_allow_html=True)
                st.dataframe(styled_df, use_container_width=True)
                st.markdown("</div>", unsafe_allow_html=True)

                # ✅ Excel download block
                from io import BytesIO
                from openpyxl.utils import get_column_letter
                from openpyxl.styles import Font
                from openpyxl.formatting.rule import ColorScaleRule

                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_results.to_excel(writer, index=False, sheet_name='Scan Results')
                    workbook = writer.book
                    worksheet = writer.sheets['Scan Results']

                    # Set column widths and font size
                    for idx, col in enumerate(df_results.columns, 1):
                        worksheet.column_dimensions[get_column_letter(idx)].width = 25
                    for row in worksheet.iter_rows():
                        for cell in row:
                            cell.font = Font(size=14)

                    # Apply color scale to Safety Score column
                    score_idx = df_results.columns.get_loc('Safety Score') + 1
                    score_letter = get_column_letter(score_idx)
                    rule = ColorScaleRule(start_type='min', start_color='FF0000', end_type='max', end_color='00FF00')
                    worksheet.conditional_formatting.add(
                        f'{score_letter}2:{score_letter}{len(df_results)+1}', rule)

                st.download_button(
                    label="Download Excel File",
                    data=output.getvalue(),
                    file_name="youtube_title_scan_results.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    except Exception as e:
        st.error(f"Something went wrong: {e}")

