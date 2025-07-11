import os
from dotenv import load_dotenv
import streamlit as st
import pandas as pd
from scan_titles_weighted_contextual_v3_riskaware import scan_titles_weighted
import requests
import logging

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

load_dotenv()

st.set_page_config(page_title="YouTube Title Scanner", layout="centered")


st.markdown(
    f"""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@300;800&display=swap');
        body {{
            background-color: #0E1117;
            color: #FFFFFF;
        }}
        .branded-content {{
            font-family: 'Montserrat', sans-serif;
        }}
        .wave-section {{
            background: linear-gradient(135deg, #951E3D, #00FFCF);
            color: #FFFFFF;
            text-align: center;
            position: relative;
            padding: 3rem 1rem 6rem;
            overflow: hidden;
        }}
        .wave-section h1 {{
            font-weight: 800;
            margin-bottom: 0.2rem;
        }}
        .wave-section .subtitle {{
            font-weight: 300;
            margin: 0;
        }}
        .wave-section svg {{
            position: absolute;
            bottom: 0;
            left: 0;
            width: 100%;
            height: 80px;
        }}
        .results-table {{
        }}
    </style>
    <div class="branded-content wave-section">
                <h1>YouTube Title Scanner</h1>
        <p class="subtitle">Scan a YouTube channel for advertiser-unfriendly words and suggest safer alternatives.</p>
        <svg viewBox="0 0 1440 100" preserveAspectRatio="none">
            <path d="M0,30 C200,80 400,0 600,30 C800,60 1000,20 1200,40 C1300,50 1400,40 1440,30 L1440,100 L0,100 Z" fill="#0E1117"></path>
        </svg>
    </div>
    """,
    unsafe_allow_html=True,
)

api_key = st.text_input(
    "Enter your YouTube Data API Key",
    value=os.getenv("YOUTUBE_API_KEY", ""),
    type="password",
)
channel_id = st.text_input("Enter the YouTube Channel ID (e.g., UC_x5XG1OV2P6uZZ5FSM9Ttw)")
max_results = st.number_input("Maximum number of titles to fetch", min_value=1, max_value=500, value=100)

def get_uploads_playlist_id(api_key, channel_id):
    url = f"https://www.googleapis.com/youtube/v3/channels?part=contentDetails&id={channel_id}&key={api_key}"
    try:
        response = requests.get(url, timeout=10)
        data = response.json()
    except requests.exceptions.RequestException as e:
        logger.error("Error fetching channel details: %s", e)
        return None
    try:
        return data['items'][0]['contentDetails']['relatedPlaylists']['uploads']
    except (KeyError, IndexError):
        return None

def fetch_video_titles(api_key, uploads_playlist_id, max_results):
    titles = []
    page_token = ""
    while len(titles) < max_results:
        url = f"https://www.googleapis.com/youtube/v3/playlistItems?key={api_key}&playlistId={uploads_playlist_id}&part=snippet&maxResults=50&pageToken={page_token}"
        try:
            response = requests.get(url, timeout=10)
            data = response.json()
        except requests.exceptions.RequestException as e:
            logger.error("Error fetching playlist items: %s", e)
            break
        for item in data.get("items", []):
            titles.append(item["snippet"]["title"])
        page_token = data.get("nextPageToken", "")
        if not page_token:
            break
    return titles[:max_results]

if st.button("Scan Titles") and api_key and channel_id:
    try:
        st.info("Fetching video titles...")
        uploads_playlist_id = get_uploads_playlist_id(api_key, channel_id)
        if not uploads_playlist_id:
            st.error("Failed to retrieve uploads playlist. Check Channel ID.")
        else:
            titles = fetch_video_titles(api_key, uploads_playlist_id, max_results)
            if not titles:
                st.warning("No titles found or API quota exceeded.")
            else:
                df_keywords = pd.read_csv("updated_keywords_expanded.csv")

                # The CSV already contains a 'keyword' column. Renaming
                # 'Flagged Keyword' blindly would create duplicate column
                # names which causes `DataFrame` objects to appear when
                # accessing `df_keywords['keyword']`. This leads to errors
                # when calling Series string methods.  Preserve a single
                # 'keyword' column instead.
                if "Flagged Keyword" in df_keywords.columns:
                    if "keyword" not in df_keywords.columns:
                        df_keywords.rename(columns={"Flagged Keyword": "keyword"}, inplace=True)
                    else:
                        df_keywords["keyword"] = df_keywords["Flagged Keyword"].fillna(df_keywords["keyword"])
                        df_keywords.drop(columns=["Flagged Keyword"], inplace=True)
                
                df_severity = pd.read_csv("safety_severity_scores.csv", encoding="utf-8-sig")

                # DEBUG: Log original column names
                logger.debug("Raw df_severity.columns = %s", df_severity.columns.tolist())

                # Clean up whitespace and stray characters
                df_severity.columns = df_severity.columns.str.strip().str.replace('"', '').str.replace("'", '')

                # DEBUG: Log cleaned column names
                logger.debug("Cleaned df_severity.columns = %s", df_severity.columns.tolist())

                # Safely rename columns if they exist
                if "Keyword" in df_severity.columns:
                    df_severity.rename(columns={"Keyword": "keyword"}, inplace=True)
                if "SeverityScoreDeduction" in df_severity.columns:
                    df_severity.rename(columns={"SeverityScoreDeduction": "severity"}, inplace=True)

                # Continue with filtering
                df_severity = df_severity[df_severity['keyword'].notna()]
                df_severity['keyword'] = df_severity['keyword'].astype(str).str.lower()

                df_results = scan_titles_weighted(titles, df_keywords, df_severity)
                # Sort results by Safety Score (ascending)
                df_results = df_results.sort_values(by="Safety Score")

                # Apply color scaling to Safety Score column
                styled_df = df_results.style.background_gradient(
                    cmap="RdYlGn", subset=["Safety Score"]
                )

                st.success("Scan complete!")
                st.markdown("<div class='results-table'>", unsafe_allow_html=True)
                st.dataframe(styled_df, use_container_width=True)
                st.markdown("</div>", unsafe_allow_html=True)

                from io import BytesIO
                from openpyxl.utils import get_column_letter
                from openpyxl.styles import Font
                from openpyxl.formatting.rule import ColorScaleRule

                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_results.to_excel(writer, index=False, sheet_name='Scan Results')
                    workbook = writer.book
                    worksheet = writer.sheets['Scan Results']

                    # Set column widths and font size
                    for idx, col in enumerate(df_results.columns, 1):
                        worksheet.column_dimensions[get_column_letter(idx)].width = 25
                    for row in worksheet.iter_rows():
                        for cell in row:
                            cell.font = Font(size=14)

                    # Apply color scale to Safety Score column
                    score_idx = df_results.columns.get_loc('Safety Score') + 1
                    score_letter = get_column_letter(score_idx)
                    rule = ColorScaleRule(start_type='min', start_color='FF0000', end_type='max', end_color='00FF00')
                    worksheet.conditional_formatting.add(
                        f'{score_letter}2:{score_letter}{len(df_results)+1}', rule)

                st.download_button(
                    label="Download Excel File",
                    data=output.getvalue(),
                    file_name="youtube_title_scan_results.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    except Exception as e:
        st.error(f"Something went wrong: {e}")
